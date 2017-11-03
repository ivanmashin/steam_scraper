"""
SteamScraper - a simple Steam scraping app.

This module contains a simple UI.
"""

__version__ = '0.1'

import tkinter as tk
import tkinter.ttk as ttk
import tkinter.filedialog
from functools import partial
import time
import steam_parser as smp
import csv
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill

import sys

DEFAULT_DAYS = '15-90'
DEFAULT_REVIEWS = 100
MAX_DAYS = 1825
MAX_REVIEWS = 100000


class Application(tk.Frame):
    indie_switch_bool = 0
    released_days_var = DEFAULT_DAYS
    reviews_var = DEFAULT_REVIEWS
    tags_var = ''
    pb_var = 0
    directory = '/'
    tags = ["Action", "Adventure", "Casual", "Massively Multiplayer",
            "Racing", "RPG", "Simulation", "Sports", "Strategy",
            "Single-player", "Multi-player", "Online Multi-Player", "Local Multi-Player",
            "Co-op", "Online Co-op", "Local Co-op", "Shared/Split Screen",
            "Demos", "Steam Workshop",
            "VR Only", "VR Supported"]
    # Headers for file saving
    info_header = ['Build date', 'Inputs', 'Amount of games', 'Average price', 'Average % of positive reviews',
                   'Owners whole', 'Players whole', 'Median playtime', 'Average playtime']
    game_data_header = ['Title', 'URL', 'Date', 'Price', '% of positive reviews', 'Reviews', 'Owners',
                        'Players Forever', 'Median forever', 'Average forever', ' ', 'Developer', 'Publisher',
                        'Owners variance', 'Players forever variance', 'Players 2 weeks', 'Players 2 weeks variance',
                        'Average 2weeks', 'Median 2weeks', 'CCU']
    gl = None
    fr = None

    def __init__(self, master=None):
        super().__init__(master)

        style = ttk.Style()
        style.configure('TButton', padding=2, width=6)

        Application.indie_switch_bool = tk.IntVar(value=1)

        Application.save_xlsx_bool = tk.IntVar(value=1)

        Application.released_days_var = tk.StringVar(value=DEFAULT_DAYS)
        #Application.released_days_var.trace("w", callback=None)

        Application.reviews_var = tk.StringVar(value=DEFAULT_REVIEWS)
        Application.reviews_var.trace('w', partial(self.int_entry_callback,
                                                   var=Application.reviews_var, minval=0, maxval=MAX_REVIEWS))

        Application.pb_var = tk.IntVar()

        self.grid()
        self.create_widgets()

    def create_widgets(self):
        """Creates all GUI"""
        UI_ROW = 0
        self.indie_switch = ttk.Checkbutton(text="Only indie",
                                            variable=Application.indie_switch_bool, onvalue=1, offvalue=0)
        self.indie_switch.grid(row=UI_ROW, column=0)

        self.save_switch = ttk.Checkbutton(text="Save as .xlsx",
                                           variable=Application.save_xlsx_bool, onvalue=1, offvalue=0)
        self.save_switch.grid(row=UI_ROW, column=1)

        UI_ROW += 1
        self.period_label = ttk.Label(text="Release period (days)")
        self.period_label.grid(row=UI_ROW, column=0, sticky=tk.E)
        self.period_entry = ttk.Entry(textvariable=Application.released_days_var)
        self.period_entry.grid(row=UI_ROW, column=1)

        UI_ROW += 1
        self.reviews_label = ttk.Label(text="Reviews >")
        self.reviews_label.grid(row=UI_ROW, column=0, sticky=tk.E)
        self.reviews_entry = ttk.Entry(textvariable=Application.reviews_var)
        self.reviews_entry.grid(row=UI_ROW, column=1)

        UI_ROW += 1
        self.tags_label = ttk.Label(text="Select tags:")
        self.tags_label.grid(row=UI_ROW, column=0, sticky=(tk.E, tk.N))
        self.tags_listbox = tk.Listbox(selectmode='extended', width=20, height=10)
        self.tags_listbox.grid(row=UI_ROW, column=1)
        self.scrollbar = ttk.Scrollbar(orient=tk.VERTICAL, command=self.tags_listbox.yview)
        self.scrollbar.grid(row=UI_ROW, column=1, sticky=(tk.E, tk.N, tk.S))
        self.tags_listbox.config(yscrollcommand=self.scrollbar.set)
        self.add_tags(self.tags_listbox)

        UI_ROW += 1
        self.status_label = ttk.Label(text="", font='TkDefaultFont 9 bold')
        self.status_label.grid(row=UI_ROW, column=1, sticky=tk.W)

        self.get_button = ttk.Button(text="GET", style='TButton', command=self.get_apps)
        self.get_button.grid(row=UI_ROW, column=0, sticky=tk.W)

        self.save_button = ttk.Button(text="SAVE", state=tk.DISABLED, style='TButton', command=self.save_copy)
        self.save_button.grid(row=UI_ROW, column=1, sticky=tk.E)

        UI_ROW += 1
        self.progress_bar = ttk.Progressbar(length=240, variable=Application.pb_var)
        self.progress_bar.grid(row=UI_ROW, columnspan=2)

        UI_ROW += 1
        self.quit_button = ttk.Button(text="QUIT", command=self.quit_app)
        self.quit_button.grid(row=UI_ROW, columnspan=2)

    def add_tags(self, listbox):
        """Get tags from Steam"""
        for tag in Application.tags:
            listbox.insert(tk.END, tag)

    def get_apps(self):
        """Get stats for specified games and convert it into .xlsx format.
        3 steps:
            1. get_suitable_apps(b_indie, period_range, reviews_num, tags_list, gui) - get games from Steam
            2. get_app_stats(appid) - stats for game from SteamSpy
            3. parse_to_xlsx(apps_data) - write game stats in Excel format
        """
        def clear():
            print("invalid release period input")
            self.status_label.config(text="Invalid input", foreground='red')
            Application.released_days_var.set("")

        # Period entry check
        period = self.period_entry.get().split('-')
        for p in period:
            if (not p.isdigit()):
                clear()
                return
        if len(period) > 1:
            cond = (len(period) > 2) \
                    or (int(period[0]) > int(period[1])) \
                    or (int(period[0]) not in range(0, MAX_DAYS)) \
                    or (int(period[1]) not in range(0, MAX_DAYS))
            if cond:
                clear()
                return
        elif (len(period) == 1) and (period[0] != ''):
            if (int(period[0]) not in range(0, MAX_DAYS)):
                clear()
                return
        print(period)

        # Check if entries isn`t empty
        reviews = self.reviews_entry.get()
        if (not period) or (not reviews):
            print("fields is empty!")
            self.status_label.config(text="Invalid input", foreground='red')
            return

        b_indie = Application.indie_switch_bool
        selected_tags = self.tags_listbox.curselection()
        tags = []
        for i in selected_tags:
            tags.append(self.tags_listbox.get(i))
        print(tags)
        print("GET")

        # Call functions here
        Application.gl, Application.fr = smp.get_suitable_apps(b_indie, period, tags, int(reviews), gui=self)
        self.progress_bar.step()
        self.status_label.config(text="Completed", foreground='black')
        self.save_button.state(['!disabled'])

    def save(self):
        initfile = 'steamscrap_' + time.strftime('%d-%m-%Y_%H-%M')
        ftypes = (('xlsx files', '*.xlsx'), ('all files', '*.*'))
        # This func returns the opened file
        f = tk.filedialog.asksaveasfile(initialdir=Application.directory, defaultextension='.xlsx',
                                        filetypes=ftypes, initialfile=initfile)
        if f is None:
            return

        # Keep last dir
        last_dir_id = f.name.rfind('/')
        Application.directory = f.name[0:last_dir_id]
        print(Application.directory)

        f.close()

    def save_copy(self):
        initfile = 'steamscrap_' + time.strftime('%d-%m-%Y_%H-%M')
        # Switch
        b_save_xlsx = Application.save_xlsx_bool.get()
        if (b_save_xlsx):
            ftypes = (('xlsx files', '*.xlsx'), ('all files', '*.*'))
            extension = '.xlsx'
        else:
            ftypes = (('csv files', '*.csv'), ('all files', '*.*'))
            extension = '.csv'

        f = tk.filedialog.asksaveasfile(initialdir=Application.directory, defaultextension=extension,
                                        filetypes=ftypes, initialfile=initfile)
        if f is None:
            return

        # Keep last dir
        last_dir_id = f.name.rfind('/')
        Application.directory = f.name[0:last_dir_id+1]
        print(Application.directory)

        # ---TEMP---
        lst1 = ['date', 'inputs']
        game_lst = [['game1', 'developer', 'price'],
                    ['game2', 'developer', 'price'],
                    ['game3', 'developer', 'price']]
        # ---TEMP---
        if (b_save_xlsx):
            Application.save_xlsx(self, filename=Application.directory+initfile, 
                                  table_info=Application.fr, games_list=Application.gl)
        else:
            Application.save_csv(self, filename=Application.directory+initfile, 
                                 table_info=Application.fr, games_list=Application.gl)
        f.close()

    def save_xlsx(self, filename, table_info, games_list):
        """
        pass filename without extension (.xlsx)
        """
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "ScrapList"

        # Cell style
        border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
        font = Font(name='Arial', size=11)
        fill = PatternFill("solid", fgColor="007FFF")

        # Create headers
        ws1.append(Application.info_header)
        ws1.append(table_info)
        ws1.append([])
        ws1.append(Application.game_data_header)
        header_rows = 4
        for col_iter in range(1, len(Application.game_data_header)+1):
            ws1.cell(row=header_rows, column=col_iter).fill = fill

        # Add games info
        for game in games_list:
            ws1.append(game)  # Write info in row (pass list)

        # Apply style
        num_rows = header_rows + len(games_list) + 1
        num_cols = len(Application.game_data_header)
        for row_iter in range(header_rows, num_rows):
            for col_iter in range(1, num_cols+1):
                cell = ws1.cell(row=row_iter, column=col_iter)
                cell.border = border
                cell.font = font
        print("xlsx save dir: ", filename+'.xlsx')
        wb.save(filename=filename+'.xlsx')

    def save_csv(self, filename, table_info, games_list):
        with open(filename+'.csv', 'w', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile, delimiter='\t', lineterminator='\n')
            writer.writerow(Application.info_header)
            writer.writerow(table_info)
            writer.writerow('')
            writer.writerow(Application.game_data_header)
            for item in games_list:
                try:
                    writer.writerow(item)
                except:
                    print(sys.exc_info())
                    print(item)

    def progress(self, number):
        self.progress_bar.step(number)
        self.update()

    def int_entry_callback(*args, var, minval, maxval):
        """Entry checker for int values in specified range"""
        value = var.get()
        if not value:
            print("empty")
            return
        if (not value.isdigit()) or (int(value) > maxval) or (int(value) < minval):
            var.set("")

    def quit_app(self):
        root.destroy()


def main():
    global root
    root = tk.Tk()
    root.title("Steam scraper")
    root.geometry('242x310')
    #root.iconbitmap('icon.ico')
    app = Application(master=root)
    app.mainloop()


if __name__ == '__main__':
    main()

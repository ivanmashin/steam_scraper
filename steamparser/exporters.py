import csv
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, PatternFill
import json


# Headers for file saving
INFO_HEADER = ['Build date', 'Inputs', 'Amount of games', 'Average price', 'Average % of positive reviews',
                'Owners whole', 'Players whole', 'Median playtime', 'Average playtime']
GAME_DATA_HEADER = ['Title', 'URL', 'Date', '$Price', '% of positive reviews', 'Reviews', 'Owners',
                    'Players Forever', 'Median forever', 'Average forever', ' ', 'Developer', 'Publisher',
                    'Owners variance', 'Players forever variance', 'Players 2 weeks', 'Players 2 weeks variance',
                    'Average 2weeks', 'Median 2weeks', 'CCU']


def save_xlsx(filename, table_info, games_list):
    """
    pass filename without extension (.xlsx)
    """
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "ScrapList"

    # Cell style
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))
    font = Font(name='Arial', size=11)
    fill = PatternFill("solid", fgColor="007FFF")

    # Create headers
    ws1.append(INFO_HEADER)
    ws1.append(table_info)
    ws1.append([])
    ws1.append(GAME_DATA_HEADER)
    header_rows = 4
    for col_iter in range(1, len(GAME_DATA_HEADER)+1):
        ws1.cell(row=header_rows, column=col_iter).fill = fill

    # Add games info
    for game in games_list:
        ws1.append(game)  # Write info in row (pass list)

    # Apply style
    num_rows = header_rows + len(games_list) + 1
    num_cols = len(GAME_DATA_HEADER)
    for row_iter in range(header_rows, num_rows):
        for col_iter in range(1, num_cols+1):
            cell = ws1.cell(row=row_iter, column=col_iter)
            cell.border = border
            cell.font = font
    print("xlsx save dir: ", filename+'.xlsx')
    wb.save(filename=filename+'.xlsx')


def save_csv(filename, table_info, games_list, separator):
    with open(filename+'.csv', 'w', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile, delimiter=separator, lineterminator='\n')
        writer.writerow(INFO_HEADER)
        writer.writerow(table_info)
        writer.writerow('')
        writer.writerow(GAME_DATA_HEADER)
        for item in games_list:
            try:
                writer.writerow(item)
            except:
                print(item)


#TODO:
def save_json(filename):
    print('WIP')
